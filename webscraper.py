# selenium 4
import os

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.firefox import GeckoDriverManager

driver = webdriver.Firefox(
    service=FirefoxService(
        GeckoDriverManager().install()
    )
)


def add_data(i,col_desc,col_data):
    # populate first row with column headers
    new_sheet.cell(row=1, column=i+1).value = col_desc.text 

    # populate second row with data
    new_sheet.cell(row=highest + 1, column=i+1).value = col_data.text

    # save to the excel
    source_wb.save("./data.xlsx")

# create excel workbook (grab exisitng or create one)
file_path = r'./data.xlsx'

if os.path.isfile(file_path):
    source_wb = openpyxl.load_workbook("./data.xlsx", data_only=True)
else:
    source_wb = openpyxl.Workbook()

# get the active sheet in each workbook
source_sheet = source_wb.active


# new_sheet = source_wb.create_sheet("New Sheet")

userInput = input("what would you like to search for from ANU data commons? ")

print("....loading")

# search using user input and rest
url = ("https://datacommons.anu.edu.au/DataCommons"
       f"/rest/search/?q={userInput}&limit=200")

driver.get(url)

driver.implicitly_wait(1)

# select articles from search
articles = driver.find_elements(By.TAG_NAME, value="article")


# loop through returned search items
for element in articles:

    driver.implicitly_wait(1)

    # find the link in the article
    link = element.find_element(By.TAG_NAME, value="a")

    driver.implicitly_wait(1)

    # open the article in a new tab
    link.send_keys(Keys.CONTROL + Keys.RETURN)

    # remove from the loop list so that when we navigate
    # back loop goes to next article
    driver.execute_script("document.querySelector('article')"
                          ".style = 'display:none;'")

    # switch to the new tab
    driver.switch_to.window(driver.window_handles[-1])

    driver.implicitly_wait(1)

    data = driver.find_element(By.CLASS_NAME, value="mt-3")
    rows = data.find_elements(By.CLASS_NAME, value="row")

    # grab the highest row with data in it
    highest = source_sheet.max_row 

    for i, row in enumerate(rows):

        col_desc = row.find_element(By.CLASS_NAME, value="col-sm-4")
        col_data = row.find_element(By.CLASS_NAME, value="col-sm-8")

        # print(row.text)
        print(col_data.text)

     # depending on article type: print to different worksheet in the workbook? double check if worksheet exists

        if col_desc.text in source_wb.sheetnames:

            #append data to that sheet
            add_data(i,col_desc,col_data)
        else:

            # if worksheet doesn't exist create one based on the catalogue type then add values to that sheet

            if (col_desc.text == "Type"):
                new_sheet = source_wb.create_sheet(f"{col_data.text}")

                #append data to that sheet
        add_data(i,col_desc,col_data)
                

    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    #  save the new workbook
    source_wb.save("./data.xlsx")
